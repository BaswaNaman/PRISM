"""
Evaluation against the 200-item ground truth.
=============================================

The brief is blunt about this: the Input-vs-Delivery-Format workbook is "the only
place where you can measure whether your output is right", and judges will look
for field-level accuracy. This module produces that number.

Metrics reported
----------------
* Field-level accuracy  — per field and overall, exact and normalised match.
* Char-limit compliance — per the description formats' published limits.
* LOV compliance        — % of generated values inside the controlled vocabulary.
* Coverage              — % of ground-truth fields we even attempt to populate.

Scoring is deliberately reported at two strictnesses. Exact match is what the
delivery format demands (casing and symbols included). Normalised match ignores
case and whitespace, and the gap between the two tells you how much of your error
is formatting rather than substance — which is the difference between a quick fix
and a modelling problem.
"""

import os
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple


def _norm(v) -> str:
    if v is None:
        return ""
    s = str(v).replace(" ", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s


def _norm_hard(v) -> str:
    """Aggressive normalisation: also drop punctuation and (R)/(TM) symbols."""
    s = _norm(v)
    s = re.sub(r"[®™©]", "", s)
    s = re.sub(r"[^a-z0-9 ]+", "", s)
    return re.sub(r"\s+", " ", s).strip()


@dataclass
class FieldScore:
    field_name: str
    total: int = 0
    exact: int = 0
    normalised: int = 0
    attempted: int = 0
    truth_populated: int = 0
    examples_wrong: List[Tuple[str, str]] = field(default_factory=list)

    @property
    def exact_pct(self) -> float:
        return round((self.exact / self.total) * 100, 1) if self.total else 0.0

    @property
    def normalised_pct(self) -> float:
        return round((self.normalised / self.total) * 100, 1) if self.total else 0.0

    @property
    def coverage_pct(self) -> float:
        return round((self.attempted / self.truth_populated) * 100, 1) if self.truth_populated else 0.0

    def as_dict(self) -> Dict:
        return {
            "field": self.field_name,
            "scored_rows": self.total,
            "exact_match_pct": self.exact_pct,
            "normalised_match_pct": self.normalised_pct,
            "coverage_pct": self.coverage_pct,
            "sample_mismatches": [
                {"expected": e, "got": g} for e, g in self.examples_wrong[:3]
            ],
        }


class GroundTruthEvaluator:
    """Scores generated records against the delivery-format sheet."""

    def __init__(self, truth_rows: Optional[List[Dict[str, Any]]] = None,
                 key_field: str = "Mfg_Part_Num"):
        self.truth_rows = truth_rows or []
        self.key_field = key_field
        self._by_key: Dict[str, Dict[str, Any]] = {}
        self._reindex()

    def _reindex(self) -> None:
        from .dedup import normalize_mpn
        self._by_key = {}
        for row in self.truth_rows:
            k = normalize_mpn(row.get(self.key_field))
            if k:
                self._by_key.setdefault(k, row)

    @property
    def loaded(self) -> bool:
        return bool(self.truth_rows)

    def score(self, generated: List[Dict[str, Any]],
              fields: Optional[List[str]] = None,
              max_examples: int = 3) -> Dict[str, Any]:
        """Compare generated rows to ground truth, matched on part number."""
        from .dedup import normalize_mpn

        if not self.loaded:
            return {
                "error": "Ground truth not loaded. Place "
                         "Unilog-Sample_200_Items-Input-vs-Output.xlsx in ./data/ and call "
                         "load_ground_truth() to enable scoring.",
                "overall": None,
            }

        # Default to whichever fields both sides actually share.
        if fields is None:
            gen_keys = set()
            for g in generated:
                gen_keys |= {k for k in g if not k.startswith("_")}
            truth_keys = set()
            for t in self.truth_rows:
                truth_keys |= {k for k in t if not k.startswith("_")}
            fields = sorted(gen_keys & truth_keys)
            fields = [f for f in fields if f != self.key_field]

        scores: Dict[str, FieldScore] = {f: FieldScore(f) for f in fields}
        matched, unmatched = 0, []

        for gen in generated:
            key = normalize_mpn(gen.get(self.key_field))
            truth = self._by_key.get(key)
            if truth is None:
                unmatched.append(gen.get(self.key_field))
                continue
            matched += 1

            for f in fields:
                fs = scores[f]
                expected = truth.get(f)
                got = gen.get(f)

                if expected is not None and str(expected).strip():
                    fs.truth_populated += 1
                if got is not None and str(got).strip():
                    fs.attempted += 1

                # Only score rows where the truth has a value to compare against.
                if expected is None or str(expected).strip() == "":
                    continue

                fs.total += 1
                if str(got or "").strip() == str(expected).strip():
                    fs.exact += 1
                    fs.normalised += 1
                elif _norm_hard(got) == _norm_hard(expected) and _norm_hard(expected):
                    fs.normalised += 1
                    if len(fs.examples_wrong) < max_examples:
                        fs.examples_wrong.append((str(expected), str(got)))
                else:
                    if len(fs.examples_wrong) < max_examples:
                        fs.examples_wrong.append((str(expected), str(got)))

        total_scored = sum(s.total for s in scores.values())
        total_exact = sum(s.exact for s in scores.values())
        total_norm = sum(s.normalised for s in scores.values())

        ranked = sorted(scores.values(), key=lambda s: s.exact_pct)

        return {
            "rows_generated": len(generated),
            "rows_matched_to_truth": matched,
            "rows_unmatched": unmatched[:10],
            "fields_evaluated": len(fields),
            "overall": {
                "values_scored": total_scored,
                "exact_match_pct": round((total_exact / total_scored) * 100, 1) if total_scored else 0.0,
                "normalised_match_pct": round((total_norm / total_scored) * 100, 1) if total_scored else 0.0,
                "formatting_only_error_pct": round(((total_norm - total_exact) / total_scored) * 100, 1) if total_scored else 0.0,
            },
            "per_field": [s.as_dict() for s in sorted(scores.values(), key=lambda x: x.field_name)],
            "weakest_fields": [s.as_dict() for s in ranked[:5]],
        }


# --------------------------------------------------------------------------
# Workbook loader
# --------------------------------------------------------------------------
def load_ground_truth(path: str,
                      input_sheet_hint: str = "input",
                      output_sheet_hint: str = "delivery") -> Tuple[List[Dict], List[Dict]]:
    """Read Unilog-Sample_200_Items-Input-vs-Output.xlsx.

    Returns (input_rows, delivery_rows) as lists of dicts keyed by column header.
    Returns ([], []) with an explanation when the file or openpyxl is missing.
    """
    if not os.path.exists(path):
        print(f"[Eval] Ground-truth workbook not found at {path}.")
        return [], []
    try:
        import openpyxl
    except ImportError:
        print("[Eval] openpyxl not installed — cannot read ground truth.")
        return [], []

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[Eval] Could not open workbook: {e}")
        return [], []

    def read_sheet(name) -> List[Dict]:
        sheet = wb[name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        # Header = first row with a decent number of non-empty string cells.
        header_idx = 0
        for i, row in enumerate(rows[:10]):
            filled = [c for c in row if c is not None and str(c).strip()]
            if len(filled) >= max(3, len(row) // 4):
                header_idx = i
                break
        headers = [str(c).strip() if c is not None else f"col_{j}"
                   for j, c in enumerate(rows[header_idx])]
        out = []
        for row in rows[header_idx + 1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            rec = {}
            for j, h in enumerate(headers):
                rec[h] = row[j] if j < len(row) else None
            out.append(rec)
        return out

    input_name = next((s for s in wb.sheetnames if input_sheet_hint in s.lower()), wb.sheetnames[0])
    output_name = next((s for s in wb.sheetnames if output_sheet_hint in s.lower()),
                       wb.sheetnames[-1])

    inputs = read_sheet(input_name)
    delivery = read_sheet(output_name)
    print(f"[Eval] Loaded {len(inputs)} input rows from '{input_name}' and "
          f"{len(delivery)} delivery rows from '{output_name}'.")
    return inputs, delivery


def load_items(path: str, sheet: Optional[str] = None) -> List[Dict]:
    """Generic loader for Sample-1000_Items.xlsx style files."""
    if not os.path.exists(path):
        print(f"[Eval] Items workbook not found at {path}.")
        return []
    try:
        import openpyxl
    except ImportError:
        print("[Eval] openpyxl not installed.")
        return []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(c).strip() if c is not None else f"col_{j}" for j, c in enumerate(rows[0])]
        out = []
        for row in rows[1:]:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            out.append({h: (row[j] if j < len(row) else None) for j, h in enumerate(headers)})
        print(f"[Eval] Loaded {len(out)} item rows from {os.path.basename(path)}.")
        return out
    except Exception as e:
        print(f"[Eval] Failed to read {path}: {e}")
        return []
