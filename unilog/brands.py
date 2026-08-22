"""
Manufacturer / brand canonicalisation against UniCat_Manufacturer_and_Brand_List.
=================================================================================

The brief requires that manufacturer and brand names match the approved list
*exactly* — legal casing, spacing, suffixes (Inc / LLC / Ltd) and (R) / (TM)
symbols included. Supplier data does not cooperate: the same manufacturer shows
up under six spellings, and empty fields arrive as literal placeholder strings.

Strategy
--------
1. Drop placeholders ("-- Unbranded --" et al.) to real None. The brief is
   explicit that these are not data.
2. Exact match on a normalised key (case/punctuation/suffix-insensitive).
3. Fall back to fuzzy match via difflib, but only accept above a confidence
   threshold. Below it the value is returned unresolved and flagged for human
   review rather than being snapped to a plausible-looking wrong company.

That last point matters: mapping "AB Controls" onto "ABB Inc." because they look
similar would be exactly the kind of confident-but-wrong output this project is
built to refuse.
"""

import os
import re
import difflib
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

# Literal placeholder strings that mean "empty" in the source data.
PLACEHOLDERS = {
    "-- unbranded --", "unbranded", "-- no unilog brand --", "no unilog brand",
    "-- no dib brand --", "no dib brand", "n/a", "na", "none", "null", "-",
    "--", "unknown", "not applicable", "no brand", "generic",
}

# Corporate suffixes ignored when comparing names.
SUFFIXES = [
    "incorporated", "inc", "corporation", "corp", "company", "co",
    "limited", "ltd", "llc", "lp", "llp", "plc", "gmbh", "ag", "sa", "nv",
    "bv", "pty", "pvt", "private", "holdings", "group", "international",
    "industries", "industrial", "manufacturing", "mfg", "products", "brands",
]

_SYMBOL_RE = re.compile(r"[®™©®™]")
_NONALNUM_RE = re.compile(r"[^a-z0-9 ]+")
_WS_RE = re.compile(r"\s+")

# Accept a fuzzy match only at or above this ratio.
FUZZY_ACCEPT_THRESHOLD = 0.88
# Between review and accept, surface as a suggestion needing confirmation.
FUZZY_REVIEW_THRESHOLD = 0.72


def is_placeholder(value) -> bool:
    """True when the string is one of the source's 'empty' markers."""
    if value is None:
        return True
    s = str(value).strip().lower()
    return s == "" or s in PLACEHOLDERS


def normalize_key(name) -> str:
    """Comparison key: lowercase, symbols and punctuation removed, suffixes dropped."""
    if name is None:
        return ""
    s = _SYMBOL_RE.sub(" ", str(name)).lower()
    s = _NONALNUM_RE.sub(" ", s)
    s = _WS_RE.sub(" ", s).strip()
    tokens = [t for t in s.split(" ") if t]
    while tokens and tokens[-1] in SUFFIXES:
        tokens.pop()
    return " ".join(tokens)


@dataclass
class BrandMatch:
    """Result of a canonicalisation attempt, carrying its own confidence."""
    input_value: Optional[str]
    manufacturer_name: Optional[str] = None
    manufacturer_code: Optional[str] = None
    brand_name: Optional[str] = None
    brand_code: Optional[str] = None
    confidence: float = 0.0
    method: str = "unresolved"      # placeholder | exact | fuzzy | unresolved
    needs_review: bool = True
    note: str = ""

    def as_dict(self) -> Dict:
        return {
            "input": self.input_value,
            "manufacturer_name": self.manufacturer_name,
            "manufacturer_code": self.manufacturer_code,
            "brand_name": self.brand_name,
            "brand_code": self.brand_code,
            "confidence": round(self.confidence, 3),
            "method": self.method,
            "needs_review": self.needs_review,
            "note": self.note,
        }


class BrandResolver:
    """Canonicalises messy supplier strings against the approved UniCat list."""

    def __init__(self, rows: Optional[List[Dict[str, str]]] = None):
        """`rows` = [{MANUFACTURER_NAME, MANUFACTURER_CODE, BRAND_NAME, BRAND_CODE}]"""
        self.rows: List[Dict[str, str]] = rows or []
        self._by_mfg_key: Dict[str, Dict[str, str]] = {}
        self._by_brand_key: Dict[str, Dict[str, str]] = {}
        self._all_keys: List[str] = []
        self._reindex()

    def _reindex(self) -> None:
        self._by_mfg_key.clear()
        self._by_brand_key.clear()
        for row in self.rows:
            mk = normalize_key(row.get("MANUFACTURER_NAME"))
            bk = normalize_key(row.get("BRAND_NAME"))
            if mk and mk not in self._by_mfg_key:
                self._by_mfg_key[mk] = row
            if bk and bk not in self._by_brand_key:
                self._by_brand_key[bk] = row
        self._all_keys = sorted(set(list(self._by_mfg_key) + list(self._by_brand_key)))

    @property
    def loaded(self) -> bool:
        return bool(self.rows)

    def resolve(self, value) -> BrandMatch:
        """Map a raw supplier string to the approved manufacturer/brand pair."""
        if is_placeholder(value):
            return BrandMatch(
                input_value=None, confidence=1.0, method="placeholder",
                needs_review=False,
                note="Source value is a placeholder ('-- Unbranded --' style) and is treated as empty, per the brief.",
            )

        raw = str(value).strip()

        if not self.loaded:
            return BrandMatch(
                input_value=raw, confidence=0.0, method="unresolved", needs_review=True,
                note="Approved manufacturer/brand list not loaded — cannot canonicalise. "
                     "Place UniCat_Manufacturer_and_Brand_List.xlsx in ./data/ to enable.",
            )

        key = normalize_key(raw)

        row = self._by_mfg_key.get(key) or self._by_brand_key.get(key)
        if row:
            return self._match_from_row(raw, row, 1.0, "exact",
                                        needs_review=False,
                                        note="Exact match on approved list after normalising case, symbols and corporate suffix.")

        close = difflib.get_close_matches(key, self._all_keys, n=1, cutoff=FUZZY_REVIEW_THRESHOLD)
        if close:
            best_key = close[0]
            score = difflib.SequenceMatcher(None, key, best_key).ratio()
            row = self._by_mfg_key.get(best_key) or self._by_brand_key.get(best_key)
            if row and score >= FUZZY_ACCEPT_THRESHOLD:
                return self._match_from_row(
                    raw, row, score, "fuzzy", needs_review=False,
                    note=f"Fuzzy matched at {score:.0%} similarity (>= {FUZZY_ACCEPT_THRESHOLD:.0%} auto-accept threshold).")
            if row:
                m = self._match_from_row(
                    raw, row, score, "fuzzy", needs_review=True,
                    note=f"Nearest approved entry is only {score:.0%} similar — below the "
                         f"{FUZZY_ACCEPT_THRESHOLD:.0%} auto-accept threshold. Suggested, not applied; needs human confirmation.")
                return m

        return BrandMatch(
            input_value=raw, confidence=0.0, method="unresolved", needs_review=True,
            note="No entry on the approved manufacturer/brand list resembles this value. Left unresolved rather than guessed.",
        )

    @staticmethod
    def _match_from_row(raw: str, row: Dict[str, str], score: float, method: str,
                        needs_review: bool, note: str) -> BrandMatch:
        mfg = (row.get("MANUFACTURER_NAME") or "").strip() or None
        brand = (row.get("BRAND_NAME") or "").strip() or None
        # Per the guidelines: where an item has no brand, the manufacturer name is used.
        return BrandMatch(
            input_value=raw,
            manufacturer_name=mfg,
            manufacturer_code=(row.get("MANUFACTURER_CODE") or "").strip() or None,
            brand_name=brand or mfg,
            brand_code=(row.get("BRAND_CODE") or "").strip() or None,
            confidence=score, method=method, needs_review=needs_review, note=note,
        )


# --------------------------------------------------------------------------
# Workbook loader
# --------------------------------------------------------------------------
def load_brand_list(path: str) -> List[Dict[str, str]]:
    """Read UniCat_Manufacturer_and_Brand_List.xlsx into row dicts.

    Returns [] with an explanatory print when the file or openpyxl is missing.
    """
    if not os.path.exists(path):
        print(f"[Brands] Workbook not found at {path} — resolver will run unloaded.")
        return []
    try:
        import openpyxl
    except ImportError:
        print("[Brands] openpyxl not installed — resolver will run unloaded.")
        return []

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        rows_iter = sheet.iter_rows(values_only=True)

        header = None
        for row in rows_iter:
            if row and any(c and "manufacturer" in str(c).lower() for c in row):
                header = [str(c).strip().upper().replace(" ", "_") if c else "" for c in row]
                break
        if header is None:
            print("[Brands] Could not locate a header row — resolver unloaded.")
            return []

        def col(*names):
            for i, h in enumerate(header):
                if any(n in h for n in names):
                    return i
            return None

        i_mfg = col("MANUFACTURER_NAME", "MANUFACTURER")
        i_mfg_code = col("MANUFACTURER_CODE", "MFG_CODE")
        i_brand = col("BRAND_NAME", "BRAND")
        i_brand_code = col("BRAND_CODE")

        out: List[Dict[str, str]] = []
        for row in rows_iter:
            if not row:
                continue

            def get(idx):
                if idx is None or idx >= len(row) or row[idx] is None:
                    return ""
                return str(row[idx]).strip()

            rec = {
                "MANUFACTURER_NAME": get(i_mfg),
                "MANUFACTURER_CODE": get(i_mfg_code),
                "BRAND_NAME": get(i_brand),
                "BRAND_CODE": get(i_brand_code),
            }
            if rec["MANUFACTURER_NAME"] or rec["BRAND_NAME"]:
                out.append(rec)

        print(f"[Brands] Loaded {len(out)} approved manufacturer/brand rows.")
        return out
    except Exception as e:
        print(f"[Brands] Failed to read workbook ({e}) — resolver unloaded.")
        return []
