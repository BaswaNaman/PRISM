"""
Controlled-vocabulary (List Of Values) validation.
=================================================

The single hardest constraint in the brief: attribute values may not be free
text. For a given classpath, only the attributes listed in the LOV apply, and
each may only take one of its permitted normalised values. As the guide puts it,
"a fluent description made of invented values scores zero."

This module turns that from a hope into a check. Every generated attribute value
is looked up; anything outside the vocabulary is reported as a violation with the
nearest legal alternative suggested. That yields "% of values found in the LOV",
one of the metrics the brief tells judges to look for.

Supports Unicat_Lov (cross-category) and the per-category specs
(FAUCETS_LOV.xlsx, Fittings_LOV.xlsx), which share the same essential shape:
    Classpath | Attribute Label | Attribute Values | Normalized Label |
    Normalized Values | Filtering Y/N
"""

import os
import re
import difflib
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Set, Tuple

SUGGEST_CUTOFF = 0.7


def _norm(s) -> str:
    """Loose comparison key for vocabulary lookup."""
    if s is None:
        return ""
    return re.sub(r"[^a-z0-9]+", " ", str(s).lower()).strip()


@dataclass
class AttributeSpec:
    """One attribute permitted on a classpath, with its controlled vocabulary."""
    label: str
    normalized_label: Optional[str] = None
    is_filterable: bool = False
    permitted_values: Set[str] = field(default_factory=set)
    # loose key -> canonical (correctly cased) value
    _value_index: Dict[str, str] = field(default_factory=dict)
    sequence: Optional[int] = None
    guidelines: Optional[str] = None

    def index(self) -> None:
        self._value_index = {_norm(v): v for v in self.permitted_values if str(v).strip()}

    def canonical(self, value) -> Optional[str]:
        """Return the approved spelling of `value`, or None if not permitted."""
        if not self._value_index:
            self.index()
        return self._value_index.get(_norm(value))

    def suggest(self, value) -> Optional[str]:
        """Closest permitted value, for a human-review hint."""
        if not self._value_index:
            self.index()
        keys = list(self._value_index)
        if not keys:
            return None
        close = difflib.get_close_matches(_norm(value), keys, n=1, cutoff=SUGGEST_CUTOFF)
        return self._value_index[close[0]] if close else None

    @property
    def is_open_vocabulary(self) -> bool:
        """True when the LOV lists no enumerated values (free-measurement fields)."""
        return not self.permitted_values


@dataclass
class ValidationIssue:
    attribute: str
    value: Optional[str]
    problem: str
    suggestion: Optional[str] = None

    def as_dict(self) -> Dict:
        return {
            "attribute": self.attribute,
            "value": self.value,
            "problem": self.problem,
            "suggestion": self.suggestion,
        }


@dataclass
class ValidationResult:
    classpath: Optional[str]
    checked: int = 0
    in_vocabulary: int = 0
    corrected: Dict[str, str] = field(default_factory=dict)
    issues: List[ValidationIssue] = field(default_factory=list)
    skipped_open: int = 0

    @property
    def lov_compliance_pct(self) -> float:
        """The judging metric: % of enumerated values that are legal LOV values."""
        return round((self.in_vocabulary / self.checked) * 100, 1) if self.checked else 0.0

    def as_dict(self) -> Dict:
        return {
            "classpath": self.classpath,
            "values_checked": self.checked,
            "values_in_lov": self.in_vocabulary,
            "lov_compliance_pct": self.lov_compliance_pct,
            "open_vocabulary_skipped": self.skipped_open,
            "corrected_to_approved_spelling": self.corrected,
            "violations": [i.as_dict() for i in self.issues],
        }


class LOVRegistry:
    """Holds classpath -> {attribute label -> AttributeSpec}."""

    def __init__(self):
        self.by_classpath: Dict[str, Dict[str, AttributeSpec]] = {}

    @property
    def loaded(self) -> bool:
        return bool(self.by_classpath)

    def classpaths(self) -> List[str]:
        return sorted(self.by_classpath)

    def add(self, classpath: str, spec: AttributeSpec) -> None:
        cp = self.by_classpath.setdefault(classpath, {})
        key = _norm(spec.label)
        if key in cp:
            cp[key].permitted_values |= spec.permitted_values
            cp[key].index()
        else:
            spec.index()
            cp[key] = spec

    def attributes_for(self, classpath: str) -> Dict[str, AttributeSpec]:
        """Attributes applicable to a classpath, with a prefix fallback so a more
        specific supplied path still resolves to its governing spec."""
        if classpath in self.by_classpath:
            return self.by_classpath[classpath]
        target = _norm(classpath)
        for cp, attrs in self.by_classpath.items():
            n = _norm(cp)
            if n and (target.startswith(n) or n.startswith(target)):
                return attrs
        return {}

    def validate(self, classpath: str, attributes: Dict[str, object]) -> ValidationResult:
        """Check every supplied attribute value against the controlled vocabulary."""
        result = ValidationResult(classpath=classpath)

        if not self.loaded:
            result.issues.append(ValidationIssue(
                attribute="(registry)", value=None,
                problem="LOV not loaded — cannot verify any value against the controlled "
                        "vocabulary. Place Unicat_Lov / <CATEGORY>_LOV.xlsx in ./data/ to enable.",
            ))
            return result

        specs = self.attributes_for(classpath)
        if not specs:
            result.issues.append(ValidationIssue(
                attribute="(classpath)", value=classpath,
                problem="Classpath not present in the LOV, so no attribute set could be resolved.",
                suggestion=self._suggest_classpath(classpath),
            ))
            return result

        for label, value in attributes.items():
            if value is None or str(value).strip() == "":
                continue
            spec = specs.get(_norm(label))

            if spec is None:
                result.checked += 1
                result.issues.append(ValidationIssue(
                    attribute=label, value=str(value),
                    problem="Attribute is not applicable to this classpath per the LOV.",
                    suggestion=self._suggest_label(label, specs),
                ))
                continue

            if spec.is_open_vocabulary:
                # Measurement-style attribute: no enumerated list to check against.
                result.skipped_open += 1
                continue

            result.checked += 1
            canon = spec.canonical(value)
            if canon is not None:
                result.in_vocabulary += 1
                if canon != str(value):
                    result.corrected[label] = canon
            else:
                result.issues.append(ValidationIssue(
                    attribute=label, value=str(value),
                    problem="Value is not in the permitted value list for this attribute.",
                    suggestion=spec.suggest(value),
                ))

        return result

    def _suggest_classpath(self, classpath: str) -> Optional[str]:
        close = difflib.get_close_matches(_norm(classpath),
                                         [_norm(c) for c in self.by_classpath],
                                         n=1, cutoff=SUGGEST_CUTOFF)
        if not close:
            return None
        for cp in self.by_classpath:
            if _norm(cp) == close[0]:
                return cp
        return None

    @staticmethod
    def _suggest_label(label: str, specs: Dict[str, AttributeSpec]) -> Optional[str]:
        close = difflib.get_close_matches(_norm(label), list(specs), n=1, cutoff=SUGGEST_CUTOFF)
        return specs[close[0]].label if close else None


# --------------------------------------------------------------------------
# Workbook loader
# --------------------------------------------------------------------------
def load_lov(path: str, registry: Optional[LOVRegistry] = None) -> LOVRegistry:
    """Load a LOV workbook (Unicat_Lov, FAUCETS_LOV, Fittings_LOV) into a registry.

    Multiple files can be merged by passing the same registry back in. Missing
    file or missing openpyxl yields an empty (clearly-unloaded) registry rather
    than an exception.
    """
    reg = registry or LOVRegistry()

    if not os.path.exists(path):
        print(f"[LOV] Workbook not found at {path} — validation will report 'not loaded'.")
        return reg
    try:
        import openpyxl
    except ImportError:
        print("[LOV] openpyxl not installed — validation will report 'not loaded'.")
        return reg

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[LOV] Could not open {path}: {e}")
        return reg

    total = 0
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        # Header can sit below title/notes rows — scan the first 20 for the row
        # that mentions an attribute column.
        header_idx, headers = None, []
        for i, row in enumerate(rows[:20]):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            if any("attribute" in c for c in cells):
                header_idx, headers = i, cells
                break
        if header_idx is None:
            continue

        def col(*names):
            for idx, h in enumerate(headers):
                if any(n in h for n in names):
                    return idx
            return None

        i_classpath = col("classpath", "class path", "leaf node", "category")
        i_label = col("attribute label", "attribute name", "attribute")
        i_norm_label = col("normalized label", "normalised label")
        i_values = col("attribute values", "permitted values", "values", "value")
        i_norm_values = col("normalized values", "normalised values")
        i_filter = col("filtering", "filterable", "filter")
        i_seq = col("sequence", "order", "seq")
        i_guide = col("guidelines", "guideline", "definition", "remarks")

        if i_label is None:
            continue

        sheet_classpath = sheet_name  # fallback when no classpath column exists

        for row in rows[header_idx + 1:]:
            if not row:
                continue

            def get(idx):
                if idx is None or idx >= len(row) or row[idx] is None:
                    return ""
                return str(row[idx]).strip()

            label = get(i_label)
            if not label:
                continue

            classpath = get(i_classpath) or sheet_classpath

            # Prefer normalised values; they are the form output must take.
            raw_values = get(i_norm_values) or get(i_values)
            values: Set[str] = set()
            if raw_values:
                for part in re.split(r"[|;\n]+|,(?![^(]*\))", raw_values):
                    p = part.strip()
                    if p and p.lower() not in ("n/a", "na", "none", "-"):
                        values.add(p)

            seq_raw = get(i_seq)
            try:
                seq = int(float(seq_raw)) if seq_raw else None
            except ValueError:
                seq = None

            spec = AttributeSpec(
                label=label,
                normalized_label=get(i_norm_label) or None,
                is_filterable=get(i_filter).upper().startswith("Y"),
                permitted_values=values,
                sequence=seq,
                guidelines=get(i_guide) or None,
            )
            reg.add(classpath, spec)
            total += 1

    print(f"[LOV] Loaded {total} attribute rows across {len(reg.by_classpath)} classpaths from {os.path.basename(path)}.")
    return reg
