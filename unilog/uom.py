"""
Unit-of-measure normalisation, per Unilog_Master_UOM_Standards.
==============================================================

Two house rules from the standard are enforced here:

1. Every unit is written in exactly one approved abbreviation form. Supplier data
   arrives as "inches", "IN.", "Inch", '"' — all of which must become "in".
2. There is always a space between the number and the unit: "24 in", never
   "24in".

The seed table below covers the measurement types that appear in the sample
data. It is intentionally a *seed*, not a claim of completeness: the client
workbook lists ~500 abbreviations across 89 measurement types. `load_uom_table()`
reads that workbook when present and merges it over these defaults, so the
authoritative list always wins. Anything not found is reported as unmapped
rather than guessed.
"""

import os
import re
from typing import Dict, Optional, Tuple, List

# --------------------------------------------------------------------------
# Seed map: variant (lowercased, punctuation-stripped) -> approved abbreviation
# --------------------------------------------------------------------------
SEED_UOM_MAP: Dict[str, str] = {
    # Length
    "in": "in", "inch": "in", "inches": "in", "\"": "in", "''": "in",
    "ft": "ft", "foot": "ft", "feet": "ft", "'": "ft",
    "mm": "mm", "millimeter": "mm", "millimetre": "mm", "millimeters": "mm",
    "cm": "cm", "centimeter": "cm", "centimetre": "cm", "centimeters": "cm",
    "m": "m", "meter": "m", "metre": "m", "meters": "m", "metres": "m",
    "yd": "yd", "yard": "yd", "yards": "yd",
    # Electrical
    "v": "V", "volt": "V", "volts": "V", "vac": "V AC", "vdc": "V DC",
    "mv": "mV", "kv": "kV",
    "a": "A", "amp": "A", "amps": "A", "ampere": "A", "amperes": "A", "amperage": "A",
    "ma": "mA", "milliamp": "mA", "milliamps": "mA",
    "w": "W", "watt": "W", "watts": "W", "kw": "kW", "kilowatt": "kW",
    "hp": "hp", "horsepower": "hp",
    "hz": "Hz", "hertz": "Hz", "khz": "kHz", "mhz": "MHz",
    "ohm": "ohm", "ohms": "ohm",
    "va": "VA", "kva": "kVA",
    # Temperature
    "c": "C", "celsius": "C", "degc": "C", "degreec": "C", "degreescelsius": "C",
    "f": "F", "fahrenheit": "F", "degf": "F", "degreef": "F",
    # Mass
    "lb": "lb", "lbs": "lb", "pound": "lb", "pounds": "lb",
    "oz": "oz", "ounce": "oz", "ounces": "oz",
    "kg": "kg", "kilogram": "kg", "kilograms": "kg", "kilo": "kg",
    "g": "g", "gram": "g", "grams": "g",
    # Pressure
    "psi": "psi", "psig": "psig", "bar": "bar", "kpa": "kPa", "mpa": "MPa",
    "inhg": "in HG", "wc": "in WC",
    # Flow
    "gpm": "gpm", "gallonsperminute": "gpm",
    "lpm": "L/min", "lmin": "L/min", "litersperminute": "L/min",
    "cfm": "cfm", "cubicfeetperminute": "cfm",
    # Rotation / sound / misc
    "rpm": "rpm", "db": "dB", "dba": "dBA", "decibel": "dB",
    "gal": "gal", "gallon": "gal", "gallons": "gal",
    "l": "L", "liter": "L", "litre": "L", "liters": "L", "litres": "L",
    "qt": "qt", "quart": "qt", "pt": "pt", "pint": "pt",
    "btu": "BTU", "btuh": "BTU/HR",
    "pc": "pc", "piece": "pc", "pieces": "pc", "ea": "ea", "each": "ea",
    "pr": "pr", "pair": "pr", "pk": "pk", "pack": "pk",
    "bx": "bx", "box": "bx", "cs": "cs", "case": "cs",
    "gauge": "ga", "ga": "ga", "awg": "AWG",
    "nominalpipesize": "NPS", "nps": "NPS",
    "mesh": "mesh", "micron": "micron", "rpmmax": "rpm",
}

_PUNCT_STRIP_RE = re.compile(r"[.\s_\-]+")


def _key(raw: str) -> str:
    """Normalise a raw unit token into a lookup key."""
    if raw is None:
        return ""
    s = str(raw).strip().lower()
    s = _PUNCT_STRIP_RE.sub("", s)
    s = s.replace("degrees", "deg").replace("°", "deg")
    if s.startswith("deg") and len(s) > 3:
        s = "deg" + s[3:]
    return s


class UOMNormalizer:
    """Normalises unit strings to the approved Unilog abbreviation."""

    def __init__(self, extra_map: Optional[Dict[str, str]] = None):
        self.map: Dict[str, str] = dict(SEED_UOM_MAP)
        if extra_map:
            # Authoritative workbook values override the seed defaults.
            for k, v in extra_map.items():
                self.map[_key(k)] = v
        self.unmapped: Dict[str, int] = {}

    def normalize_unit(self, unit: Optional[str]) -> Tuple[Optional[str], bool]:
        """Return (approved_abbreviation, was_recognised).

        When the unit is not in the approved list the original string is returned
        with was_recognised=False and the token is recorded in `self.unmapped`, so
        gaps are reportable instead of silently passing through as if compliant.
        """
        if unit is None or str(unit).strip() == "":
            return None, True
        k = _key(unit)
        if k in self.map:
            return self.map[k], True
        self.unmapped[str(unit).strip()] = self.unmapped.get(str(unit).strip(), 0) + 1
        return str(unit).strip(), False

    def format_measurement(self, value, unit: Optional[str]) -> str:
        """Render 'value unit' with the mandatory single space. 24, 'inches' -> '24 in'."""
        approved, _ = self.normalize_unit(unit)
        if value is None:
            return approved or ""
        val_str = self._clean_number(value)
        if not approved:
            return val_str
        return f"{val_str} {approved}"

    @staticmethod
    def _clean_number(value) -> str:
        """Drop a pointless trailing .0 so 24.0 renders as 24."""
        try:
            f = float(value)
            if f == int(f):
                return str(int(f))
            return str(f).rstrip("0").rstrip(".")
        except (TypeError, ValueError):
            return str(value).strip()

    # ----------------------------------------------------------------------
    # Free-text repair
    # ----------------------------------------------------------------------
    # NOTE: alternatives are ordered longest-first. Regex alternation is greedy in
    # order, so listing "a" before "amps" would match the "a" and leave "mps",
    # failing the trailing (?![A-Za-z]) guard and silently skipping the fix.
    GLUED_RE = re.compile(
        r"(?<![A-Za-z])(\d+(?:\.\d+)?|\d+-\d+/\d+|\d+/\d+)"
        r"(inches|inch|milliamps|milliamp|amperes|ampere|amps|amp|volts|volt|"
        r"kilograms|kilogram|pounds|pound|ounces|ounce|grams|gram|"
        r"kilowatt|watts|watt|hertz|khz|mhz|hz|"
        r"psig|psi|bar|kpa|mpa|gpm|lpm|cfm|rpm|dba|db|btu|gal|"
        r"lbs|lb|oz|kg|mm|cm|ft|in|vac|vdc|kw|ma|hp|kv|mv|"
        r"[wvalgm])"
        r"(?![A-Za-z])",
        re.IGNORECASE,
    )

    def fix_spacing_in_text(self, text: str) -> str:
        """Insert the required space in glued measurements: '24in' -> '24 in'.

        Also normalises the unit token itself, so '50-1/4IN' becomes '50-1/4 in'.
        """
        if not text:
            return text

        def _sub(m):
            number, unit = m.group(1), m.group(2)
            approved, ok = self.normalize_unit(unit)
            return f"{number} {approved if ok else unit}"

        return self.GLUED_RE.sub(_sub, str(text))

    def report_unmapped(self) -> List[Tuple[str, int]]:
        """Units encountered that are not in the approved list, most frequent first."""
        return sorted(self.unmapped.items(), key=lambda kv: -kv[1])


# --------------------------------------------------------------------------
# Optional: load the authoritative workbook when the client data is present
# --------------------------------------------------------------------------
def load_uom_table(path: str) -> Dict[str, str]:
    """Read Unilog_Master_UOM_Standards_Abbreviations_and_Terms.xlsx.

    Returns {variant: approved_abbreviation}. Returns {} (and prints why) when the
    file or openpyxl is unavailable, so the caller keeps working on seed defaults
    instead of crashing.
    """
    if not os.path.exists(path):
        print(f"[UOM] Workbook not found at {path} — using seed table only.")
        return {}
    try:
        import openpyxl
    except ImportError:
        print("[UOM] openpyxl not installed — using seed table only.")
        return {}

    mapping: Dict[str, str] = {}
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return {}

        # Sheets in this pack have multi-row headers; find the header row by
        # looking for the one that mentions an abbreviation-ish column.
        header_idx, headers = 0, []
        for i, row in enumerate(rows[:15]):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            if any("abbrev" in c or "capture" in c or "uom" in c for c in cells):
                header_idx, headers = i, cells
                break

        def find_col(*names):
            for idx, h in enumerate(headers):
                if any(n in h for n in names):
                    return idx
            return None

        col_approved = find_col("abbrev", "capture", "approved")
        col_variant = find_col("term", "measurement", "name", "unit", "description")

        if col_approved is None:
            print("[UOM] Could not locate an abbreviation column — using seed table.")
            return {}

        for row in rows[header_idx + 1:]:
            if not row:
                continue
            approved = row[col_approved] if col_approved < len(row) else None
            if approved is None or str(approved).strip() == "":
                continue
            approved = str(approved).strip()
            mapping[approved] = approved
            if col_variant is not None and col_variant < len(row):
                variant = row[col_variant]
                if variant is not None and str(variant).strip():
                    mapping[str(variant).strip()] = approved

        print(f"[UOM] Loaded {len(mapping)} unit variants from workbook.")
        return mapping
    except Exception as e:
        print(f"[UOM] Failed to read workbook ({e}) — using seed table only.")
        return {}


DEFAULT = UOMNormalizer()
